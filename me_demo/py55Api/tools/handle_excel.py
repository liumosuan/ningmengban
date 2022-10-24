# -*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook

class HandleExcel:
    def __init__(self, filename, sheet_name):
        # 加载excel
        self.wb_obj = load_workbook(filename=filename)  # 实例化加载excel对象
        # 获取sheet
        self.sheet_obj = self.wb_obj.worksheets[sheet_name]

    def get_excel_test_cases(self):
        # 用户存储每个字典的值
        case_list = []
        # 行切片
        datas = list(self.sheet_obj.iter_rows(values_only=True))
        case_title = datas[0]
        case_datas = datas[1:]
        for case in case_datas:
            # zip(list1,list2)  将数据转成成kv对
            result = dict(zip(case_title, case))
            case_list.append(result)
        # print(case_list)
        return case_list

    def close_excel(self):
        # 关闭excel
        self.wb_obj.close()


if __name__ == '__main__':
    data = HandleExcel(filename=r"D:\Pycharm\ningmengban\me_demo\py55Api\test_data\case_data.xlsx",
                       sheet_name=0)
    data.get_excel_test_cases()
    data.close_excel()
