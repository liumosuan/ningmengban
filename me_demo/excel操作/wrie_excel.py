# -*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook


"""
写excel
"""
# 加载excel
wb_obj = load_workbook("接口测试框架实践用例.xlsx")
# 获取sheet
sheet_obj = wb_obj.worksheets[0]
# 打印获取的sheet名称
print(sheet_obj)
# 修改单元格信息
# sheet_obj['A8'] = "test"
# sheet_obj.cell(row=7, column=1, value="demo")
# # 保存文件，要将打开的文件先关闭
# wb_obj.save("接口测试框架实践用例.xlsx")
# 关闭excel
wb_obj.close()
