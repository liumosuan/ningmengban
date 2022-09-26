"""
load_workbook()
filename:excel文件名
read_only=False:只读，不可编辑
data_only=False:只读取公式。True：读取计算结果
"""
import openpyxl
from openpyxl import load_workbook

# # 加载excel
# wb_obj = load_workbook(filename="接口测试框架实践用例.xlsx")
# # 获取表单对象
# sheet_obj = wb_obj["1.登录"]
# print(sheet_obj)
# # 获取单元格
# # cell_obj = sheet_obj['A2']    # 定位单元格
# # cell_obj = sheet_obj.cell(3,3)  # 第几行第几列
# # 获取单元格内容
# # print(cell_obj.value)
# # 关闭excel
# wb_obj.close()

# 加载excel
wb_obj = load_workbook("接口测试框架实践用例.xlsx")
# 获取表单所有名称,
# names = wb_obj.sheetnames
# print(names, type(names))

# 获取表单对象，通过索引的方式
sheet_obj = wb_obj.worksheets[0]
# 获取表单对象，通过列表取值
# sheet_obj = wb_obj["1.登录"]
# 打印第一个sheet的名称
print(sheet_obj)

# 第一页表单的最大行
rows = sheet_obj.max_row
# 第一页表单的最大列
columns = sheet_obj.max_column

# 打印最大行和最大列
# print(rows)
# print(columns)

# 获取所有的行对象,行对象的类型为元组
rows_obj = sheet_obj.rows
# print(rows_obj)

# 打印所有的行对象
# for obj in rows_obj:
#     print(obj,type(obj))

# 按行打印所有的单元格，单元格的类型为cell
# for obj in rows_obj:
#     for row in obj:
#         print(row)  # 按行打印所有单元格
# 按行打印所有单元格的值
# print(row.value)
# 打印单元格类型
# print(type(row))

"""
行切片
min_row=None,   最小行，起始行的索引值（索引从1开始，Int类型，默认1）
max_row=None,   最大行，起始行的索引值（索引从1开始，Int类型，默认最大行的值）
min_col=None,   最小列，起始列的索引值（索引从1开始，Int类型，默认1）
max_col=None,   最大列，起始列的索引值（索引从1开始，Int类型，默认最大列的值）
values_only=False   false返回单元格的对象，例如A1就是单元格的对象;true返回单元格对应的值

**切片原则是两头都包含，包含起始位置索引的值，也包含结束位置的索引值
行切片就是数据是以元组形式，每一行每一行打印
"""
# result = sheet_obj.iter_rows(min_row=1, max_row=2, min_col=1,
#                              max_col=2, values_only=True)
result = sheet_obj.iter_rows(min_row=1, max_row=3, min_col=1,
                             max_col=3, values_only=True)
# result是一个生成器
# print(result, type(result))
# print("行切片的值：", list(result))

"""
列切片
行切片就是数据是以元组形式，每一列每一列打印
"""
# result2 = sheet_obj.iter_cols(min_col=1, max_col=2, min_row=1,
#                               max_row=2, values_only=True)
result2 = sheet_obj.iter_cols(min_col=1, max_col=3, min_row=1,
                              max_row=3, values_only=True)
# print("列切片的值：", list(result2))




# 获取单元格
cell_obj = sheet_obj["A1"]
# 第二行第一列单元格的信息
# cell_obj = sheet_obj.cell(2, 3)
# 打印单元格的值
# print(f"{sheet_obj['A1']}单元格的值为:", cell_obj.value)
# print(f"{sheet_obj['A1']}单元格值的类型为:", type(cell_obj.value))
# 关闭excel
wb_obj.close()
