# -*- coding:utf-8 -*-
"""
数据封装
"""
from openpyxl.reader.excel import load_workbook


class HandleExcel:
    def __init__(self, filename, sheet_name):
        # 加载excel
        self.wb_obj = load_workbook(filename=filename)
        # 获取sheet
        self.sheet_obj = self.wb_obj.worksheets[sheet_name]

    def get_excel_test_cases(self):
        # 用于存储每对字典的值
        cases_list = []
        # 行切片
        datas = list(self.sheet_obj.iter_rows(values_only=True))
        case_title = datas[0]  # 表头
        case_datas = datas[1:]  # 数据
        # print(case_title)
        # print(case_datas)
        for case in case_datas:
            # zip(list1,list2) # 将数据转换成kv，(key,value)
            result = dict(zip(case_title, case))
            cases_list.append(result)
        print(cases_list)
        return cases_list

    def close_excel(self):
        # 关闭excel
        self.wb_obj.close()


# 单元测试
if __name__ == '__main__':
    data = HandleExcel(filename="接口测试用例.xlsx", sheet_name=0)
    data.get_excel_test_cases()
