"""
load_workbook()
filename:excel文件名
read_only=False:只读，不可编辑
data_only=False:只读取公式。True：读取计算结果
"""

from openpyxl import load_workbook

# 加载excel
wb_obj = load_workbook(filename="接口测试框架实践用例.xlsx")
# 获取表单对象
sheet_obj = wb_obj["1.登录"]
# 获取单元格
# cell_obj = sheet_obj['A2']    # 定位单元格
cell_obj = sheet_obj.cell(3,3)  # 第几行第几列
# 获取单元格内容
print(cell_obj.value)
# 关闭excel
wb_obj.close()
