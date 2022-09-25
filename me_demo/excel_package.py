# -*- coding:utf-8 -*-
"""
数据封装
"""
from openpyxl.reader.excel import load_workbook


def au():
    print("hello")
# def get_excel_test_cases():
#     # 加载excel
#     wb_obj = load_workbook(filename="接口测试框架实践用例.xlsx")
#     # 获取sheet
#     sheet_obj = wb_obj.worksheets[0]
#     # 行切片
#     datas = list(sheet_obj.iter_rows(values_only=True))
#     case_title = datas[0]  # 表头
#     case_datas = datas[1:]  # 数据
#     print(case_title)
#     print(case_datas)
#     # print(sheet_obj)
#     # 关闭excel
#     wb_obj.close()
